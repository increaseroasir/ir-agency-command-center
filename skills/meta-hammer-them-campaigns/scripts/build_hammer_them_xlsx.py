"""
Meta Ads Bulk Upload XLSX Builder - Hammer Them Campaigns
Base template: templates/meta_export_template_630cols.xlsx (630 cols, full account export)
"""

import os
import sys
import argparse
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timezone, timedelta

def build_xlsx(output_path, client_name, zip_list, included_audiences, rt_audiences, page_id):
    # Load template relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, '..', 'templates', 'meta_export_template_630cols.xlsx')
    
    wb_template = openpyxl.load_workbook(template_path)
    ws_template = wb_template.active
    headers = [cell.value for cell in ws_template[1]]

    col_idx = {h: i for i, h in enumerate(headers) if h}

    # Format ZIP codes
    ZIP_CODES = ', '.join(f'US:{z.strip()}' for z in zip_list)
    INCLUDED_AUDIENCES = ', '.join(included_audiences)

    # Times - CDT = UTC-5
    now_cdt = datetime.now(timezone(timedelta(hours=-5)))
    start_time_str = now_cdt.strftime('%m/%d/%Y %I:%M:%S %p').lower()
    end_dt = now_cdt.replace(hour=16, minute=0, second=0, microsecond=0) + timedelta(days=1)
    end_time_str = end_dt.strftime('%m/%d/%Y %I:%M:%S %p').lower()

    # Create new workbook
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = ws_template.title
    for i, h in enumerate(headers):
        ws_new.cell(row=1, column=i+1, value=h)

    def make_row():
        return [None] * len(headers)

    def set_col(row, key, val):
        if key in col_idx:
            row[col_idx[key]] = val

    for row_num, rt in enumerate(rt_audiences, start=2):
        rt_name = rt['name']
        audience_id = rt['id']
        audience_name = rt['audience_name']
        
        row = make_row()

        # --- CAMPAIGN LEVEL (blank ID = create new) ---
        set_col(row, 'Campaign Name', f'{client_name} | RT | Blitz 0-2d')
        set_col(row, 'Campaign Status', 'ACTIVE')
        set_col(row, 'Campaign Objective', 'Outcome Engagement')
        set_col(row, 'Buying Type', 'AUCTION')
        set_col(row, 'Campaign Daily Budget', '50000') # $500/day CBO
        set_col(row, 'Campaign Bid Strategy', 'Highest volume or value')
        set_col(row, 'Special Ad Categories', 'None')
        set_col(row, 'New Objective', 'Yes')
        set_col(row, 'Buy With Prime Type', 'NONE')
        set_col(row, 'Is Budget Scheduling Enabled For Campaign', 'No')
        set_col(row, 'Campaign High Demand Periods', '[]')
        set_col(row, 'Buy With Integration Partner', 'NONE')

        # --- AD SET LEVEL (blank ID = create new) ---
        set_col(row, 'Ad Set Name', f'RT | {rt_name}')
        set_col(row, 'Ad Set Run Status', 'ACTIVE')
        set_col(row, 'Ad Set Time Start', start_time_str)
        set_col(row, 'Ad Set Time Stop', end_time_str)
        set_col(row, 'Ad Set Lifetime Impressions', '0')
        set_col(row, 'Ad Set Lifetime Budget', '0')
        set_col(row, 'Use Accelerated Delivery', 'No')
        set_col(row, 'Is Budget Scheduling Enabled For Ad Set', 'No')
        set_col(row, 'Ad Set High Demand Periods', '[]')
        set_col(row, 'Use Dynamic Creative', 'No')

        # Destination type: ON_VIDEO = "On your ad > Video views"
        set_col(row, 'Destination Type', 'ON_VIDEO')

        # Page link object
        set_col(row, 'Link Object ID', f'o:{page_id}')

        # Geo - ZIP codes only
        set_col(row, 'Zip', ZIP_CODES)
        set_col(row, 'Location Types', 'home, recent')

        # Demographics
        set_col(row, 'Age Min', '25')
        set_col(row, 'Age Max', '65')

        # Audiences
        set_col(row, 'Custom Audiences', INCLUDED_AUDIENCES)
        set_col(row, 'Excluded Custom Audiences', f'{audience_id}:{audience_name}')

        # Targeting settings
        set_col(row, 'Advantage Audience', '0')
        set_col(row, 'Individual Setting', 'geo: On')
        set_col(row, 'Targeting Optimization', 'none')
        set_col(row, 'Targeting Relaxation', 'custom_audience: Off, lookalike: Off')

        # Brand safety
        set_col(row, 'Brand Safety Inventory Filtering Levels', 'FACEBOOK_RELAXED, AN_RELAXED')

        # Optimization
        set_col(row, 'Optimization Goal', 'THRUPLAY')
        set_col(row, 'Attribution Spec', '[{"event_type":"CLICK_THROUGH","window_days":1}]')
        set_col(row, 'Billing Event', 'IMPRESSIONS')
        set_col(row, 'Ad Set Bid Strategy', 'Highest volume or value')

        # Write row
        for col_i, val in enumerate(row):
            ws_new.cell(row=row_num, column=col_i+1, value=val)

    wb_new.save(output_path)
    print(f"Success! Built {len(rt_audiences)} ad sets in {output_path}")

if __name__ == '__main__':
    # This is meant to be called by Manus programmatically, but can be modified for CLI args if needed.
    print("This script is a module meant to be imported and executed by Manus.")
    print("See SKILL.md for usage instructions.")
